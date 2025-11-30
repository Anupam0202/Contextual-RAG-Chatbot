"""
Advanced Analytics Module with Enhanced Features
Implements comprehensive analytics, reporting, and Excel export functionality
"""

import json
import logging
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime, timedelta
from dataclasses import dataclass, field
from collections import defaultdict, Counter
import asyncio
from pathlib import Path
import io
import math

import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import plotly.io as pio
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart, PieChart
from openpyxl.formatting.rule import DataBarRule

logger = logging.getLogger(__name__)

# Modern color palette
COLORS = {
    'primary': '#667eea',
    'secondary': '#764ba2',
    'success': '#48bb78',
    'warning': '#f6ad55',
    'danger': '#fc8181',
    'info': '#63b3ed',
    'dark': '#2d3748',
    'light': '#f7fafc',
    'gradient': ['#667eea', '#764ba2', '#f093fb', '#f5576c', '#4facfe'],
    'chart_colors': ['#667eea', '#48bb78', '#f6ad55', '#fc8181', '#63b3ed', '#9f7aea', '#ed64a6', '#38b2ac']
}

@dataclass
class QueryMetrics:
    """Detailed metrics for each query"""
    timestamp: datetime
    session_id: str
    query: str
    response_time: float
    chunks_retrieved: int
    confidence: float
    context_size: int = 0
    response_length: int = 0
    user_satisfaction: Optional[float] = None
    topic_category: Optional[str] = None
    error_occurred: bool = False
    model_name: str = ""
    tokens_used: int = 0

@dataclass
class SessionMetrics:
    """Metrics for a user session"""
    session_id: str
    start_time: datetime
    end_time: Optional[datetime] = None
    query_count: int = 0
    total_response_time: float = 0.0
    avg_confidence: float = 0.0
    documents_used: List[str] = field(default_factory=list)
    
class AdvancedAnalytics:
    """Enhanced analytics engine with comprehensive metrics and visualizations"""
    
    def __init__(self):
        self.query_history: List[QueryMetrics] = []
        self.session_analytics: Dict[str, SessionMetrics] = {}
        self.document_analytics: Dict[str, Dict] = {}
        self.hourly_stats: Dict[int, List[float]] = defaultdict(list)
        self.daily_stats: Dict[str, Dict] = defaultdict(lambda: {'queries': 0, 'response_times': [], 'confidences': []})
        
        self.performance_thresholds = {
            'response_time': {'excellent': 1.0, 'good': 2.0, 'acceptable': 5.0},
            'confidence': {'excellent': 0.9, 'good': 0.8, 'acceptable': 0.6},
            'chunks': {'optimal_min': 3, 'optimal_max': 7}
        }
        
        self._cache = {}
        self._cache_ttl = 300  # 5 minutes
        
    def addQueryMetric(self, metric: QueryMetrics):
        """Add a query metric to history with enhanced tracking"""
        self.query_history.append(metric)
        
        # Update hourly stats
        hour = metric.timestamp.hour
        self.hourly_stats[hour].append(metric.response_time)
        
        # Update daily stats
        date_key = metric.timestamp.strftime('%Y-%m-%d')
        self.daily_stats[date_key]['queries'] += 1
        self.daily_stats[date_key]['response_times'].append(metric.response_time)
        self.daily_stats[date_key]['confidences'].append(metric.confidence)
        
        # Update session analytics
        if metric.session_id not in self.session_analytics:
            self.session_analytics[metric.session_id] = SessionMetrics(
                session_id=metric.session_id,
                start_time=metric.timestamp
            )
        
        session = self.session_analytics[metric.session_id]
        session.end_time = metric.timestamp
        session.query_count += 1
        session.total_response_time += metric.response_time
        session.avg_confidence = (session.avg_confidence * (session.query_count - 1) + metric.confidence) / session.query_count
        
        # Invalidate cache
        self._cache.clear()
    
    def getMetricsDataFrame(self) -> pd.DataFrame:
        """Convert query history to DataFrame for analysis"""
        if not self.query_history:
            return pd.DataFrame()
        
        data = []
        for q in self.query_history:
            data.append({
                'timestamp': q.timestamp,
                'date': q.timestamp.date(),
                'hour': q.timestamp.hour,
                'day_of_week': q.timestamp.strftime('%A'),
                'session_id': q.session_id,
                'query': q.query,
                'query_length': len(q.query),
                'response_time': q.response_time,
                'chunks_retrieved': q.chunks_retrieved,
                'confidence': q.confidence,
                'context_size': q.context_size,
                'response_length': q.response_length,
                'error_occurred': q.error_occurred,
                'tokens_used': q.tokens_used
            })
        
        return pd.DataFrame(data)
    
    def generateInteractiveReport(self) -> Dict[str, Any]:
        """Generate comprehensive interactive analytics report"""
        try:
            if not self.query_history or len(self.query_history) < 1:
                return self._getEmptyReportStructure()
            
            # Check cache
            cache_key = f"report_{len(self.query_history)}"
            if cache_key in self._cache:
                cached_time, cached_data = self._cache[cache_key]
                if datetime.now() - cached_time < timedelta(seconds=self._cache_ttl):
                    return cached_data
            
            # Generate report components
            summary = self._generateExecutiveSummary()
            insights = self._generateInsights()
            recommendations = self._generateRecommendations()
            visualizations = self._createVisualizations()
            predictions = self._generatePredictions()
            trends = self._analyzeTrends()
            
            report = {
                'generated_at': datetime.now().isoformat(),
                'summary': summary,
                'insights': insights,
                'recommendations': recommendations,
                'visualizations': visualizations,
                'predictions': predictions,
                'trends': trends,
                'detailed_metrics': self._getDetailedMetrics(),
                'performance_breakdown': self._getPerformanceBreakdown()
            }
            
            # Cache the report
            self._cache[cache_key] = (datetime.now(), report)
            
            return report
            
        except Exception as e:
            logger.error(f"Error generating report: {e}")
            return self._getEmptyReportStructure()
    
    def _getEmptyReportStructure(self) -> Dict[str, Any]:
        """Return empty report structure with styled placeholder charts"""
        empty_fig = self._createStyledEmptyChart("Start chatting to generate analytics!")
        
        return {
            'generated_at': datetime.now().isoformat(),
            'summary': {
                'total_queries': 0,
                'avg_response_time': 0,
                'success_rate': 100,
                'user_satisfaction_score': 0,
                'total_sessions': 0,
                'avg_session_duration': 0,
                'avg_confidence': 0,
                'total_chunks_processed': 0,
                'queries_per_session': 0,
                'peak_hour': 'N/A',
                'busiest_day': 'N/A',
                'total_tokens': 0
            },
            'insights': ["📊 Upload documents and start chatting to generate insights"],
            'recommendations': ["🚀 Begin using the chatbot to receive personalized recommendations"],
            'visualizations': {
                'overview_dashboard': empty_fig,
                'time_series': empty_fig,
                'performance_gauges': empty_fig,
                'engagement_heatmap': empty_fig,
                'topic_distribution': empty_fig,
                'confidence_histogram': empty_fig,
                'response_time_distribution': empty_fig,
                'session_analysis': empty_fig,
                'trend_analysis': empty_fig
            },
            'predictions': {'message': 'Insufficient data for predictions'},
            'trends': {},
            'detailed_metrics': {},
            'performance_breakdown': {}
        }
    
    def _createStyledEmptyChart(self, message: str) -> go.Figure:
        """Create styled empty chart with gradient background"""
        fig = go.Figure()
        
        # Add decorative elements
        fig.add_shape(
            type="circle",
            xref="paper", yref="paper",
            x0=0.3, y0=0.3, x1=0.7, y1=0.7,
            line=dict(color=COLORS['primary'], width=3, dash="dash"),
            opacity=0.3
        )
        
        fig.add_annotation(
            text=f"📊<br><br>{message}",
            xref="paper", yref="paper",
            x=0.5, y=0.5,
            showarrow=False,
            font=dict(size=18, color=COLORS['dark']),
            align="center"
        )
        
        fig.update_layout(
            height=400,
            showlegend=False,
            xaxis=dict(showgrid=False, showticklabels=False, zeroline=False),
            yaxis=dict(showgrid=False, showticklabels=False, zeroline=False),
            plot_bgcolor='rgba(0,0,0,0)',
            paper_bgcolor='rgba(0,0,0,0)',
            margin=dict(l=20, r=20, t=20, b=20)
        )
        
        return fig
    
    def _generateExecutiveSummary(self) -> Dict[str, Any]:
        """Generate comprehensive executive summary statistics"""
        if not self.query_history:
            return {}
        
        df = self.getMetricsDataFrame()
        
        # Basic stats
        total_queries = len(self.query_history)
        avg_response_time = df['response_time'].mean()
        median_response_time = df['response_time'].median()
        p95_response_time = df['response_time'].quantile(0.95)
        
        # Success rate
        error_count = df['error_occurred'].sum()
        success_rate = ((total_queries - error_count) / total_queries) * 100
        
        # Confidence stats
        avg_confidence = df['confidence'].mean()
        min_confidence = df['confidence'].min()
        max_confidence = df['confidence'].max()
        
        # Session stats
        total_sessions = len(self.session_analytics)
        avg_queries_per_session = total_queries / max(total_sessions, 1)
        
        # Time-based stats
        peak_hour = df.groupby('hour').size().idxmax() if len(df) > 0 else 0
        busiest_day = df.groupby('day_of_week').size().idxmax() if len(df) > 0 else 'N/A'
        
        # Calculate user satisfaction score (composite metric)
        satisfaction_score = self._calculateSatisfactionScore(df)
        
        # Performance grade
        performance_grade = self._calculatePerformanceGrade(avg_response_time, avg_confidence, success_rate)
        
        # Trends
        recent_queries = df[df['timestamp'] > datetime.now() - timedelta(hours=1)]
        queries_last_hour = len(recent_queries)
        
        return {
            'total_queries': total_queries,
            'avg_response_time': round(avg_response_time, 3),
            'median_response_time': round(median_response_time, 3),
            'p95_response_time': round(p95_response_time, 3),
            'success_rate': round(success_rate, 1),
            'user_satisfaction_score': round(satisfaction_score, 1),
            'performance_grade': performance_grade,
            'total_sessions': total_sessions,
            'avg_session_duration': self._calculateAvgSessionDuration(),
            'avg_confidence': round(avg_confidence, 3),
            'min_confidence': round(min_confidence, 3),
            'max_confidence': round(max_confidence, 3),
            'total_chunks_processed': int(df['chunks_retrieved'].sum()),
            'avg_chunks_per_query': round(df['chunks_retrieved'].mean(), 1),
            'queries_per_session': round(avg_queries_per_session, 1),
            'peak_hour': f"{peak_hour}:00",
            'busiest_day': busiest_day,
            'queries_last_hour': queries_last_hour,
            'total_tokens': int(df['tokens_used'].sum()) if 'tokens_used' in df.columns else 0,
            'avg_query_length': round(df['query_length'].mean(), 0),
            'avg_response_length': round(df['response_length'].mean(), 0)
        }
    
    def _calculateSatisfactionScore(self, df: pd.DataFrame) -> float:
        """Calculate composite user satisfaction score (0-100)"""
        if len(df) == 0:
            return 0
        
        # Response time component (30% weight)
        avg_response = df['response_time'].mean()
        if avg_response < self.performance_thresholds['response_time']['excellent']:
            rt_score = 100
        elif avg_response < self.performance_thresholds['response_time']['good']:
            rt_score = 85
        elif avg_response < self.performance_thresholds['response_time']['acceptable']:
            rt_score = 70
        else:
            rt_score = max(50, 100 - (avg_response - 5) * 5)
        
        # Confidence component (40% weight)
        avg_confidence = df['confidence'].mean()
        conf_score = avg_confidence * 100
        
        # Error rate component (20% weight)
        error_rate = df['error_occurred'].mean()
        error_score = (1 - error_rate) * 100
        
        # Consistency component (10% weight) - lower variance is better
        response_time_std = df['response_time'].std()
        consistency_score = max(0, 100 - response_time_std * 20)
        
        # Weighted average
        satisfaction = (
            rt_score * 0.30 +
            conf_score * 0.40 +
            error_score * 0.20 +
            consistency_score * 0.10
        )
        
        return min(100, max(0, satisfaction))
    
    def _calculatePerformanceGrade(self, avg_response: float, avg_confidence: float, success_rate: float) -> str:
        """Calculate overall performance grade (A+ to F)"""
        score = 0
        
        # Response time scoring
        if avg_response < 1:
            score += 35
        elif avg_response < 2:
            score += 30
        elif avg_response < 3:
            score += 25
        elif avg_response < 5:
            score += 20
        else:
            score += 10
        
        # Confidence scoring
        score += avg_confidence * 35
        
        # Success rate scoring
        score += (success_rate / 100) * 30
        
        # Grade mapping
        if score >= 95:
            return 'A+'
        elif score >= 90:
            return 'A'
        elif score >= 85:
            return 'A-'
        elif score >= 80:
            return 'B+'
        elif score >= 75:
            return 'B'
        elif score >= 70:
            return 'B-'
        elif score >= 65:
            return 'C+'
        elif score >= 60:
            return 'C'
        elif score >= 55:
            return 'C-'
        elif score >= 50:
            return 'D'
        else:
            return 'F'
    
    def _calculateAvgSessionDuration(self) -> float:
        """Calculate average session duration in minutes"""
        if not self.session_analytics:
            return 0
        
        durations = []
        for session in self.session_analytics.values():
            if session.start_time and session.end_time:
                duration = (session.end_time - session.start_time).total_seconds() / 60
                if duration > 0:
                    durations.append(duration)
        
        return round(sum(durations) / len(durations), 1) if durations else 0
    
    def _generateInsights(self) -> List[str]:
        """Generate AI-powered insights from analytics data"""
        insights = []
        
        if not self.query_history:
            return ["📊 Insufficient data for insights"]
        
        df = self.getMetricsDataFrame()
        
        # Performance insights
        avg_response = df['response_time'].mean()
        if avg_response < 1:
            insights.append("⚡ **Excellent Performance**: Average response time is under 1 second!")
        elif avg_response < 2:
            insights.append("✅ **Good Performance**: Response times are within optimal range")
        elif avg_response > 5:
            insights.append("⚠️ **Performance Alert**: Response times are higher than recommended")
        
        # Peak usage insights
        hourly_counts = df.groupby('hour').size()
        if len(hourly_counts) > 0:
            peak_hour = hourly_counts.idxmax()
            peak_count = hourly_counts.max()
            insights.append(f"📈 **Peak Usage**: Most queries occur at {peak_hour}:00 ({peak_count} queries)")
        
        # Confidence insights
        avg_confidence = df['confidence'].mean()
        low_confidence_pct = (df['confidence'] < 0.6).mean() * 100
        
        if avg_confidence >= 0.85:
            insights.append(f"🎯 **High Confidence**: {avg_confidence:.1%} average confidence score")
        elif low_confidence_pct > 20:
            insights.append(f"📚 **Knowledge Gap**: {low_confidence_pct:.0f}% of queries have low confidence - consider adding more documents")
        
        # Trend insights
        if len(df) >= 10:
            recent = df.tail(5)['response_time'].mean()
            older = df.head(5)['response_time'].mean()
            
            if recent < older * 0.8:
                insights.append("📉 **Improving**: Response times have decreased by 20%+ recently")
            elif recent > older * 1.2:
                insights.append("📈 **Trending Up**: Response times have increased recently")
        
        # Session insights
        if len(self.session_analytics) > 0:
            avg_queries = len(self.query_history) / len(self.session_analytics)
            if avg_queries > 10:
                insights.append(f"💬 **High Engagement**: Average {avg_queries:.0f} queries per session")
        
        # Error insights
        error_rate = df['error_occurred'].mean() * 100
        if error_rate == 0:
            insights.append("✨ **Zero Errors**: Perfect success rate!")
        elif error_rate > 5:
            insights.append(f"🔧 **Error Rate**: {error_rate:.1f}% of queries encountered errors")
        
        # Query pattern insights
        if 'query_length' in df.columns:
            avg_query_len = df['query_length'].mean()
            if avg_query_len > 100:
                insights.append("📝 **Detailed Queries**: Users are asking comprehensive questions")
            elif avg_query_len < 30:
                insights.append("💡 **Tip**: Longer, more specific queries often yield better results")
        
        return insights if insights else ["📊 Analyzing patterns..."]
    
    def _generateRecommendations(self) -> List[str]:
        """Generate actionable recommendations"""
        recommendations = []
        
        if not self.query_history:
            return ["🚀 Start using the chatbot to receive personalized recommendations"]
        
        df = self.getMetricsDataFrame()
        
        # Performance recommendations
        avg_response = df['response_time'].mean()
        if avg_response > 3:
            recommendations.append("⚡ **Speed Up**: Enable caching in settings to improve response times")
            recommendations.append("📊 **Optimize Retrieval**: Reduce chunk count or use smaller chunk sizes")
        
        # Confidence recommendations
        avg_confidence = df['confidence'].mean()
        if avg_confidence < 0.7:
            recommendations.append("📚 **Improve Quality**: Upload more relevant documents to boost answer accuracy")
            recommendations.append("🔧 **Tune Settings**: Adjust similarity threshold in retrieval settings")
        
        # Chunk optimization
        avg_chunks = df['chunks_retrieved'].mean()
        if avg_chunks > 8:
            recommendations.append("📦 **Chunk Optimization**: Consider reducing retrieval top-k for faster responses")
        elif avg_chunks < 3:
            recommendations.append("📈 **More Context**: Increase retrieval top-k for more comprehensive answers")
        
        # Usage recommendations
        if len(self.session_analytics) > 0:
            avg_duration = self._calculateAvgSessionDuration()
            if avg_duration < 2:
                recommendations.append("💡 **Explore More**: Try asking follow-up questions for deeper insights")
        
        # Error recommendations
        error_rate = df['error_occurred'].mean()
        if error_rate > 0.05:
            recommendations.append("🔍 **Reduce Errors**: Check document formatting and query complexity")
        
        return recommendations if recommendations else ["✅ System performing optimally - no recommendations needed"]
    
    def _createVisualizations(self) -> Dict[str, go.Figure]:
        """Create comprehensive interactive Plotly visualizations"""
        visualizations = {}
        
        try:
            if not self.query_history or len(self.query_history) < 1:
                empty_fig = self._createStyledEmptyChart("No data available")
                return {key: empty_fig for key in [
                    'overview_dashboard', 'time_series', 'performance_gauges',
                    'engagement_heatmap', 'topic_distribution', 'confidence_histogram',
                    'response_time_distribution', 'session_analysis', 'trend_analysis'
                ]}
            
            df = self.getMetricsDataFrame()
            
            # 1. Overview Dashboard (Multi-metric)
            visualizations['overview_dashboard'] = self._createOverviewDashboard(df)
            
            # 2. Time Series Analysis
            visualizations['time_series'] = self._createTimeSeriesChart(df)
            
            # 3. Performance Gauges
            visualizations['performance_gauges'] = self._createPerformanceGauges(df)
            
            # 4. Engagement Heatmap
            visualizations['engagement_heatmap'] = self._createEngagementHeatmap(df)
            
            # 5. Topic Distribution
            visualizations['topic_distribution'] = self._createTopicDistribution(df)
            
            # 6. Confidence Histogram
            visualizations['confidence_histogram'] = self._createConfidenceHistogram(df)
            
            # 7. Response Time Distribution
            visualizations['response_time_distribution'] = self._createResponseTimeDistribution(df)
            
            # 8. Session Analysis
            visualizations['session_analysis'] = self._createSessionAnalysis(df)
            
            # 9. Trend Analysis
            visualizations['trend_analysis'] = self._createTrendAnalysis(df)
            
        except Exception as e:
            logger.error(f"Error creating visualizations: {e}")
            empty_fig = self._createStyledEmptyChart("Error creating visualization")
            for key in ['overview_dashboard', 'time_series', 'performance_gauges',
                       'engagement_heatmap', 'topic_distribution', 'confidence_histogram',
                       'response_time_distribution', 'session_analysis', 'trend_analysis']:
                if key not in visualizations:
                    visualizations[key] = empty_fig
        
        return visualizations
    
    def _createOverviewDashboard(self, df: pd.DataFrame) -> go.Figure:
        """Create multi-metric overview dashboard"""
        fig = make_subplots(
            rows=2, cols=3,
            specs=[
                [{"type": "indicator"}, {"type": "indicator"}, {"type": "indicator"}],
                [{"type": "bar"}, {"type": "scatter"}, {"type": "pie"}]
            ],
            subplot_titles=('', '', '', 'Queries by Hour', 'Response Time Trend', 'Query Types'),
            vertical_spacing=0.15,
            horizontal_spacing=0.1
        )
        
        # Row 1: Key Metrics as Indicators
        avg_response = df['response_time'].mean()
        avg_confidence = df['confidence'].mean()
        total_queries = len(df)
        
        # Response Time Indicator
        fig.add_trace(
            go.Indicator(
                mode="gauge+number",
                value=avg_response,
                title={'text': "Avg Response Time (s)", 'font': {'size': 14}},
                gauge={
                    'axis': {'range': [0, 10], 'tickwidth': 1},
                    'bar': {'color': COLORS['primary']},
                    'bgcolor': "white",
                    'borderwidth': 2,
                    'bordercolor': "gray",
                    'steps': [
                        {'range': [0, 2], 'color': COLORS['success']},
                        {'range': [2, 5], 'color': COLORS['warning']},
                        {'range': [5, 10], 'color': COLORS['danger']}
                    ],
                    'threshold': {
                        'line': {'color': "red", 'width': 4},
                        'thickness': 0.75,
                        'value': 5
                    }
                },
                number={'suffix': 's', 'font': {'size': 24}}
            ),
            row=1, col=1
        )
        
        # Confidence Indicator
        fig.add_trace(
            go.Indicator(
                mode="gauge+number",
                value=avg_confidence * 100,
                title={'text': "Avg Confidence (%)", 'font': {'size': 14}},
                gauge={
                    'axis': {'range': [0, 100], 'tickwidth': 1},
                    'bar': {'color': COLORS['secondary']},
                    'bgcolor': "white",
                    'borderwidth': 2,
                    'bordercolor': "gray",
                    'steps': [
                        {'range': [0, 60], 'color': COLORS['danger']},
                        {'range': [60, 80], 'color': COLORS['warning']},
                        {'range': [80, 100], 'color': COLORS['success']}
                    ]
                },
                number={'suffix': '%', 'font': {'size': 24}}
            ),
            row=1, col=2
        )
        
        # Total Queries Indicator
        fig.add_trace(
            go.Indicator(
                mode="number+delta",
                value=total_queries,
                title={'text': "Total Queries", 'font': {'size': 14}},
                delta={'reference': max(1, total_queries - len(df[df['timestamp'] > datetime.now() - timedelta(hours=1)])), 'relative': True},
                number={'font': {'size': 36, 'color': COLORS['primary']}}
            ),
            row=1, col=3
        )
        
        # Row 2: Charts
        # Hourly distribution bar chart
        hourly_counts = df.groupby('hour').size().reindex(range(24), fill_value=0)
        fig.add_trace(
            go.Bar(
                x=list(range(24)),
                y=hourly_counts.values,
                marker_color=COLORS['gradient'][0],
                name='Queries'
            ),
            row=2, col=1
        )
        
        # Response time trend
        if len(df) > 1:
            fig.add_trace(
                go.Scatter(
                    x=df['timestamp'],
                    y=df['response_time'],
                    mode='lines+markers',
                    line=dict(color=COLORS['primary'], width=2),
                    marker=dict(size=6),
                    name='Response Time'
                ),
                row=2, col=2
            )
        
        # Query type distribution
        query_types = self._categorizeQueries(df['query'].tolist())
        fig.add_trace(
            go.Pie(
                labels=list(query_types.keys()),
                values=list(query_types.values()),
                hole=0.4,
                marker_colors=COLORS['chart_colors']
            ),
            row=2, col=3
        )
        
        fig.update_layout(
            height=600,
            showlegend=False,
            title_text="📊 Analytics Overview Dashboard",
            title_x=0.5,
            title_font_size=20,
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)'
        )
        
        return fig
    
    def _createTimeSeriesChart(self, df: pd.DataFrame) -> go.Figure:
        """Create comprehensive time series analysis chart"""
        if len(df) < 2:
            return self._createStyledEmptyChart("Need more data for time series")
        
        fig = make_subplots(
            rows=3, cols=1,
            subplot_titles=('Query Volume Over Time', 'Response Time Trend', 'Confidence Score Trend'),
            vertical_spacing=0.1,
            shared_xaxes=True
        )
        
        # Resample data by appropriate interval
        df_sorted = df.sort_values('timestamp')
        
        # Determine resampling frequency based on data span
        time_span = (df_sorted['timestamp'].max() - df_sorted['timestamp'].min()).total_seconds()
        if time_span < 3600:  # Less than 1 hour
            resample_freq = '5T'  # 5 minutes
        elif time_span < 86400:  # Less than 1 day
            resample_freq = '1H'  # 1 hour
        else:
            resample_freq = '1D'  # 1 day
        
        df_resampled = df_sorted.set_index('timestamp')
        
        # Query volume
        query_counts = df_resampled.resample(resample_freq).size()
        fig.add_trace(
            go.Scatter(
                x=query_counts.index,
                y=query_counts.values,
                mode='lines+markers',
                fill='tozeroy',
                line=dict(color=COLORS['primary'], width=2),
                marker=dict(size=8),
                name='Queries',
                fillcolor=f"rgba(102, 126, 234, 0.3)"
            ),
            row=1, col=1
        )
        
        # Response time with moving average
        response_time = df_resampled['response_time'].resample(resample_freq).mean()
        fig.add_trace(
            go.Scatter(
                x=response_time.index,
                y=response_time.values,
                mode='lines+markers',
                line=dict(color=COLORS['warning'], width=2),
                marker=dict(size=8),
                name='Avg Response Time'
            ),
            row=2, col=1
        )
        
        # Add threshold line
        fig.add_hline(y=2, line_dash="dash", line_color="red", 
                     annotation_text="Target", row=2, col=1)
        
        # Confidence trend
        confidence = df_resampled['confidence'].resample(resample_freq).mean()
        fig.add_trace(
            go.Scatter(
                x=confidence.index,
                y=confidence.values,
                mode='lines+markers',
                line=dict(color=COLORS['success'], width=2),
                marker=dict(size=8),
                name='Avg Confidence'
            ),
            row=3, col=1
        )
        
        fig.update_layout(
            height=700,
            showlegend=True,
            title_text="📈 Time Series Analysis",
            title_x=0.5,
            title_font_size=20,
            hovermode='x unified',
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)'
        )
        
        fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='rgba(128,128,128,0.2)')
        fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='rgba(128,128,128,0.2)')
        
        return fig
    
    def _createPerformanceGauges(self, df: pd.DataFrame) -> go.Figure:
        """Create performance metrics gauge dashboard"""
        fig = make_subplots(
            rows=1, cols=4,
            specs=[[{'type': 'indicator'}, {'type': 'indicator'}, {'type': 'indicator'}, {'type': 'indicator'}]],
            subplot_titles=('Response Time', 'Confidence', 'Success Rate', 'Efficiency')
        )
        
        avg_response = df['response_time'].mean()
        avg_confidence = df['confidence'].mean()
        success_rate = (1 - df['error_occurred'].mean()) * 100
        
        # Calculate efficiency (composite metric)
        efficiency = (avg_confidence * 100 + (10 - min(avg_response, 10)) * 10 + success_rate) / 3
        
        # Response Time Gauge
        fig.add_trace(
            go.Indicator(
                mode="gauge+number+delta",
                value=avg_response,
                delta={'reference': 2, 'decreasing': {'color': 'green'}, 'increasing': {'color': 'red'}},
                gauge={
                    'axis': {'range': [0, 10]},
                    'bar': {'color': COLORS['primary']},
                    'steps': [
                        {'range': [0, 2], 'color': 'rgba(72, 187, 120, 0.3)'},
                        {'range': [2, 5], 'color': 'rgba(246, 173, 85, 0.3)'},
                        {'range': [5, 10], 'color': 'rgba(252, 129, 129, 0.3)'}
                    ],
                    'threshold': {'line': {'color': "red", 'width': 4}, 'thickness': 0.75, 'value': 5}
                },
                number={'suffix': 's', 'valueformat': '.2f'}
            ),
            row=1, col=1
        )
        
        # Confidence Gauge
        fig.add_trace(
            go.Indicator(
                mode="gauge+number+delta",
                value=avg_confidence * 100,
                delta={'reference': 80},
                gauge={
                    'axis': {'range': [0, 100]},
                    'bar': {'color': COLORS['secondary']},
                    'steps': [
                        {'range': [0, 60], 'color': 'rgba(252, 129, 129, 0.3)'},
                        {'range': [60, 80], 'color': 'rgba(246, 173, 85, 0.3)'},
                        {'range': [80, 100], 'color': 'rgba(72, 187, 120, 0.3)'}
                    ]
                },
                number={'suffix': '%', 'valueformat': '.1f'}
            ),
            row=1, col=2
        )
        
        # Success Rate Gauge
        fig.add_trace(
            go.Indicator(
                mode="gauge+number",
                value=success_rate,
                gauge={
                    'axis': {'range': [0, 100]},
                    'bar': {'color': COLORS['success']},
                    'steps': [
                        {'range': [0, 90], 'color': 'rgba(246, 173, 85, 0.3)'},
                        {'range': [90, 100], 'color': 'rgba(72, 187, 120, 0.3)'}
                    ]
                },
                number={'suffix': '%', 'valueformat': '.1f'}
            ),
            row=1, col=3
        )
        
        # Efficiency Gauge
        fig.add_trace(
            go.Indicator(
                mode="gauge+number",
                value=efficiency,
                gauge={
                    'axis': {'range': [0, 100]},
                    'bar': {'color': COLORS['info']},
                    'steps': [
                        {'range': [0, 50], 'color': 'rgba(252, 129, 129, 0.3)'},
                        {'range': [50, 75], 'color': 'rgba(246, 173, 85, 0.3)'},
                        {'range': [75, 100], 'color': 'rgba(72, 187, 120, 0.3)'}
                    ]
                },
                number={'suffix': '%', 'valueformat': '.1f'}
            ),
            row=1, col=4
        )
        
        fig.update_layout(
            height=300,
            title_text="🎯 Performance Metrics",
            title_x=0.5,
            title_font_size=20,
            paper_bgcolor='rgba(0,0,0,0)'
        )
        
        return fig
    
    def _createEngagementHeatmap(self, df: pd.DataFrame) -> go.Figure:
        """Create user engagement heatmap by hour and day"""
        if len(df) < 2:
            return self._createStyledEmptyChart("Need more data for heatmap")
        
        # Create pivot table
        df['day_num'] = df['timestamp'].dt.dayofweek
        
        heatmap_data = df.pivot_table(
            values='response_time',
            index='day_of_week',
            columns='hour',
            aggfunc='count',
            fill_value=0
        )
        
        # Ensure all hours are present
        for hour in range(24):
            if hour not in heatmap_data.columns:
                heatmap_data[hour] = 0
        
        heatmap_data = heatmap_data.reindex(columns=range(24), fill_value=0)
        
        # Sort by day of week
        day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        existing_days = [day for day in day_order if day in heatmap_data.index]
        if existing_days:
            heatmap_data = heatmap_data.reindex(existing_days, fill_value=0)
        
        fig = go.Figure(data=go.Heatmap(
            z=heatmap_data.values,
            x=[f'{h:02d}:00' for h in range(24)],
            y=heatmap_data.index.tolist(),
            colorscale=[
                [0, 'rgba(102, 126, 234, 0.1)'],
                [0.25, 'rgba(102, 126, 234, 0.3)'],
                [0.5, 'rgba(118, 75, 162, 0.5)'],
                [0.75, 'rgba(118, 75, 162, 0.7)'],
                [1, 'rgba(118, 75, 162, 1)']
            ],
            text=heatmap_data.values,
            texttemplate='%{text}',
            textfont={"size": 10},
            hovertemplate='Day: %{y}<br>Hour: %{x}<br>Queries: %{z}<extra></extra>',
            colorbar=dict(title="Queries", tickformat='d')
        ))
        
        fig.update_layout(
            title_text="🔥 User Engagement Heatmap",
            title_x=0.5,
            title_font_size=20,
            xaxis_title="Hour of Day",
            yaxis_title="Day of Week",
            height=400,
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)'
        )
        
        return fig
    
    def _createTopicDistribution(self, df: pd.DataFrame) -> go.Figure:
        """Create topic distribution pie/sunburst chart"""
        query_types = self._categorizeQueries(df['query'].tolist())
        
        if not query_types:
            return self._createStyledEmptyChart("No topic data available")
        
        # Create sunburst chart
        labels = list(query_types.keys())
        values = list(query_types.values())
        
        fig = go.Figure(data=[
            go.Pie(
                labels=labels,
                values=values,
                hole=0.4,
                marker=dict(
                    colors=COLORS['chart_colors'][:len(labels)],
                    line=dict(color='white', width=2)
                ),
                textposition='inside',
                textinfo='percent+label',
                hovertemplate='<b>%{label}</b><br>Count: %{value}<br>Percentage: %{percent}<extra></extra>'
            )
        ])
        
        fig.update_layout(
            title_text="💭 Query Topic Distribution",
            title_x=0.5,
            title_font_size=20,
            height=400,
            annotations=[dict(text='Topics', x=0.5, y=0.5, font_size=16, showarrow=False)],
            paper_bgcolor='rgba(0,0,0,0)'
        )
        
        return fig
    
    def _createConfidenceHistogram(self, df: pd.DataFrame) -> go.Figure:
        """Create confidence score distribution histogram"""
        fig = go.Figure()
        
        # Histogram
        fig.add_trace(go.Histogram(
            x=df['confidence'],
            nbinsx=20,
            name='Frequency',
            marker=dict(
                color=COLORS['primary'],
                line=dict(color='white', width=1)
            ),
            opacity=0.8
        ))
        
        # Add threshold lines
        thresholds = [
            (0.6, 'Acceptable', 'orange'),
            (0.8, 'Good', 'yellow'),
            (0.9, 'Excellent', 'green')
        ]
        
        for threshold, label, color in thresholds:
            fig.add_vline(
                x=threshold,
                line_dash="dash",
                line_color=color,
                annotation_text=label,
                annotation_position="top"
            )
        
        # Add KDE curve
        from scipy import stats
        if len(df['confidence']) > 5:
            kde_x = np.linspace(0, 1, 100)
            kde = stats.gaussian_kde(df['confidence'].dropna())
            kde_y = kde(kde_x) * len(df) * 0.05  # Scale to histogram
            
            fig.add_trace(go.Scatter(
                x=kde_x,
                y=kde_y,
                mode='lines',
                name='Distribution',
                line=dict(color=COLORS['secondary'], width=3)
            ))
        
        # Add mean line
        mean_conf = df['confidence'].mean()
        fig.add_vline(
            x=mean_conf,
            line_dash="solid",
            line_color="red",
            annotation_text=f"Mean: {mean_conf:.2f}",
            annotation_position="top right"
        )
        
        fig.update_layout(
            title_text="🎯 Confidence Score Distribution",
            title_x=0.5,
            title_font_size=20,
            xaxis_title="Confidence Score",
            yaxis_title="Frequency",
            height=400,
            showlegend=True,
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)'
        )
        
        fig.update_xaxes(range=[0, 1], showgrid=True, gridwidth=1, gridcolor='rgba(128,128,128,0.2)')
        fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='rgba(128,128,128,0.2)')
        
        return fig
    
    def _createResponseTimeDistribution(self, df: pd.DataFrame) -> go.Figure:
        """Create response time distribution with box plot and histogram"""
        fig = make_subplots(
            rows=1, cols=2,
            subplot_titles=('Distribution', 'Box Plot'),
            column_widths=[0.7, 0.3]
        )
        
        # Histogram with gradient fill
        fig.add_trace(
            go.Histogram(
                x=df['response_time'],
                nbinsx=30,
                name='Response Time',
                marker=dict(
                    color=df['response_time'],
                    colorscale=[[0, COLORS['success']], [0.5, COLORS['warning']], [1, COLORS['danger']]],
                    line=dict(color='white', width=0.5)
                )
            ),
            row=1, col=1
        )
        
        # Box plot
        fig.add_trace(
            go.Box(
                y=df['response_time'],
                name='Distribution',
                marker_color=COLORS['primary'],
                boxpoints='outliers'
            ),
            row=1, col=2
        )
        
        # Add percentile annotations
        p50 = df['response_time'].median()
        p95 = df['response_time'].quantile(0.95)
        
        fig.add_annotation(
            x=p50, y=0,
            text=f"Median: {p50:.2f}s",
            showarrow=True,
            arrowhead=2,
            row=1, col=1
        )
        
        fig.update_layout(
            title_text="⚡ Response Time Analysis",
            title_x=0.5,
            title_font_size=20,
            height=400,
            showlegend=False,
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)'
        )
        
        fig.update_xaxes(title_text="Response Time (s)", row=1, col=1)
        fig.update_yaxes(title_text="Frequency", row=1, col=1)
        
        return fig
    
    def _createSessionAnalysis(self, df: pd.DataFrame) -> go.Figure:
        """Create session-based analysis visualization"""
        if len(self.session_analytics) < 1:
            return self._createStyledEmptyChart("No session data available")
        
        # Prepare session data
        session_data = []
        for sid, session in self.session_analytics.items():
            session_data.append({
                'session_id': sid[:8],
                'queries': session.query_count,
                'avg_response_time': session.total_response_time / max(session.query_count, 1),
                'avg_confidence': session.avg_confidence,
                'duration': (session.end_time - session.start_time).total_seconds() / 60 if session.end_time else 0
            })
        
        session_df = pd.DataFrame(session_data)
        
        fig = make_subplots(
            rows=1, cols=2,
            subplot_titles=('Queries per Session', 'Session Metrics'),
            specs=[[{'type': 'bar'}, {'type': 'scatter'}]]
        )
        
        # Bar chart of queries per session
        fig.add_trace(
            go.Bar(
                x=session_df['session_id'],
                y=session_df['queries'],
                marker_color=COLORS['primary'],
                name='Queries'
            ),
            row=1, col=1
        )
        
        # Scatter plot of response time vs confidence
        fig.add_trace(
            go.Scatter(
                x=session_df['avg_response_time'],
                y=session_df['avg_confidence'],
                mode='markers',
                marker=dict(
                    size=session_df['queries'] * 3 + 10,
                    color=session_df['queries'],
                    colorscale='Viridis',
                    showscale=True,
                    colorbar=dict(title="Queries")
                ),
                text=session_df['session_id'],
                hovertemplate='Session: %{text}<br>Response Time: %{x:.2f}s<br>Confidence: %{y:.2f}<extra></extra>',
                name='Sessions'
            ),
            row=1, col=2
        )
        
        fig.update_layout(
            title_text="👥 Session Analysis",
            title_x=0.5,
            title_font_size=20,
            height=400,
            showlegend=False,
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)'
        )
        
        fig.update_xaxes(title_text="Session", row=1, col=1)
        fig.update_yaxes(title_text="Query Count", row=1, col=1)
        fig.update_xaxes(title_text="Avg Response Time (s)", row=1, col=2)
        fig.update_yaxes(title_text="Avg Confidence", row=1, col=2)
        
        return fig
    
    def _createTrendAnalysis(self, df: pd.DataFrame) -> go.Figure:
        """Create trend analysis with moving averages"""
        if len(df) < 5:
            return self._createStyledEmptyChart("Need more data for trend analysis")
        
        df_sorted = df.sort_values('timestamp').reset_index(drop=True)
        
        # Calculate moving averages
        window = min(5, len(df_sorted))
        df_sorted['response_time_ma'] = df_sorted['response_time'].rolling(window=window, min_periods=1).mean()
        df_sorted['confidence_ma'] = df_sorted['confidence'].rolling(window=window, min_periods=1).mean()
        
        fig = make_subplots(
            rows=2, cols=1,
            subplot_titles=('Response Time Trend', 'Confidence Trend'),
            vertical_spacing=0.15
        )
        
        # Response time with MA
        fig.add_trace(
            go.Scatter(
                x=list(range(len(df_sorted))),
                y=df_sorted['response_time'],
                mode='markers',
                marker=dict(size=8, color=COLORS['primary'], opacity=0.5),
                name='Response Time'
            ),
            row=1, col=1
        )
        
        fig.add_trace(
            go.Scatter(
                x=list(range(len(df_sorted))),
                y=df_sorted['response_time_ma'],
                mode='lines',
                line=dict(color=COLORS['danger'], width=3),
                name=f'{window}-Query Moving Avg'
            ),
            row=1, col=1
        )
        
        # Confidence with MA
        fig.add_trace(
            go.Scatter(
                x=list(range(len(df_sorted))),
                y=df_sorted['confidence'],
                mode='markers',
                marker=dict(size=8, color=COLORS['success'], opacity=0.5),
                name='Confidence'
            ),
            row=2, col=1
        )
        
        fig.add_trace(
            go.Scatter(
                x=list(range(len(df_sorted))),
                y=df_sorted['confidence_ma'],
                mode='lines',
                line=dict(color=COLORS['secondary'], width=3),
                name=f'{window}-Query Moving Avg'
            ),
            row=2, col=1
        )
        
        fig.update_layout(
            title_text="📉 Trend Analysis",
            title_x=0.5,
            title_font_size=20,
            height=500,
            showlegend=True,
            paper_bgcolor='rgba(0,0,0,0)',
            plot_bgcolor='rgba(0,0,0,0)',
            hovermode='x unified'
        )
        
        fig.update_xaxes(title_text="Query Number", showgrid=True, gridwidth=1, gridcolor='rgba(128,128,128,0.2)')
        fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='rgba(128,128,128,0.2)')
        
        return fig
    
    def _categorizeQueries(self, queries: List[str]) -> Dict[str, int]:
        """Categorize queries by type"""
        categories = defaultdict(int)
        
        for query in queries:
            query_lower = query.lower()
            
            if any(word in query_lower for word in ['what', 'define', 'explain', 'describe']):
                categories['Informational'] += 1
            elif any(word in query_lower for word in ['how', 'step', 'process', 'guide']):
                categories['Procedural'] += 1
            elif any(word in query_lower for word in ['why', 'reason', 'cause', 'because']):
                categories['Analytical'] += 1
            elif any(word in query_lower for word in ['compare', 'difference', 'versus', 'vs']):
                categories['Comparative'] += 1
            elif any(word in query_lower for word in ['list', 'enumerate', 'all', 'types']):
                categories['Enumerative'] += 1
            elif any(word in query_lower for word in ['summarize', 'summary', 'brief', 'overview']):
                categories['Summary'] += 1
            else:
                categories['General'] += 1
        
        return dict(categories)
    
    def _analyzeTrends(self) -> Dict[str, Any]:
        """Analyze trends in the data"""
        if len(self.query_history) < 5:
            return {'message': 'Insufficient data for trend analysis'}
        
        df = self.getMetricsDataFrame()
        
        # Response time trend
        recent = df.tail(max(1, len(df)//3))['response_time'].mean()
        older = df.head(max(1, len(df)//3))['response_time'].mean()
        
        if older > 0:
            response_time_change = ((recent - older) / older) * 100
        else:
            response_time_change = 0
        
        # Confidence trend
        recent_conf = df.tail(max(1, len(df)//3))['confidence'].mean()
        older_conf = df.head(max(1, len(df)//3))['confidence'].mean()
        
        if older_conf > 0:
            confidence_change = ((recent_conf - older_conf) / older_conf) * 100
        else:
            confidence_change = 0
        
        return {
            'response_time_trend': 'improving' if response_time_change < -5 else 'stable' if abs(response_time_change) <= 5 else 'degrading',
            'response_time_change_pct': round(response_time_change, 1),
            'confidence_trend': 'improving' if confidence_change > 5 else 'stable' if abs(confidence_change) <= 5 else 'declining',
            'confidence_change_pct': round(confidence_change, 1),
            'query_volume_trend': self._analyzeQueryVolumeTrend(df)
        }
    
    def _analyzeQueryVolumeTrend(self, df: pd.DataFrame) -> str:
        """Analyze query volume trend"""
        if len(df) < 2:
            return 'insufficient_data'
        
        df_sorted = df.sort_values('timestamp')
        mid_point = len(df_sorted) // 2
        
        first_half_rate = mid_point / max(1, (df_sorted.iloc[mid_point]['timestamp'] - df_sorted.iloc[0]['timestamp']).total_seconds() / 3600)
        second_half_rate = (len(df_sorted) - mid_point) / max(1, (df_sorted.iloc[-1]['timestamp'] - df_sorted.iloc[mid_point]['timestamp']).total_seconds() / 3600)
        
        if second_half_rate > first_half_rate * 1.2:
            return 'increasing'
        elif second_half_rate < first_half_rate * 0.8:
            return 'decreasing'
        return 'stable'
    
    def _generatePredictions(self) -> Dict[str, Any]:
        """Generate predictive analytics"""
        if len(self.query_history) < 10:
            return {'message': 'Need at least 10 queries for predictions'}
        
        df = self.getMetricsDataFrame()
        
        # Peak hour prediction
        hourly_counts = df.groupby('hour').size()
        peak_hour = hourly_counts.idxmax() if len(hourly_counts) > 0 else 0
        
        # Expected queries next hour
        current_hour = datetime.now().hour
        historical_rate = hourly_counts.get(current_hour, 0)
        
        # Response time prediction
        recent_times = df.tail(10)['response_time'].tolist()
        if len(recent_times) >= 3:
            # Simple linear trend
            trend = (recent_times[-1] - recent_times[0]) / len(recent_times)
            predicted_time = recent_times[-1] + trend
        else:
            predicted_time = df['response_time'].mean()
        
        return {
            'predicted_peak_hour': peak_hour,
            'expected_queries_next_hour': max(1, int(historical_rate)),
            'predicted_response_time': round(max(0.1, predicted_time), 2),
            'confidence_forecast': round(df['confidence'].rolling(5).mean().iloc[-1], 2) if len(df) >= 5 else df['confidence'].mean()
        }
    
    def _getDetailedMetrics(self) -> Dict[str, Any]:
        """Get detailed metrics for export"""
        if not self.query_history:
            return {}
        
        df = self.getMetricsDataFrame()
        
        response_times = df['response_time'].tolist()
        confidences = df['confidence'].tolist()
        
        return {
            'total_queries': len(self.query_history),
            'unique_sessions': len(self.session_analytics),
            'error_rate': round(df['error_occurred'].mean() * 100, 2),
            'percentiles': {
                'response_time_p25': round(np.percentile(response_times, 25), 3),
                'response_time_p50': round(np.percentile(response_times, 50), 3),
                'response_time_p75': round(np.percentile(response_times, 75), 3),
                'response_time_p95': round(np.percentile(response_times, 95), 3),
                'response_time_p99': round(np.percentile(response_times, 99), 3),
                'confidence_p25': round(np.percentile(confidences, 25), 3),
                'confidence_p50': round(np.percentile(confidences, 50), 3),
                'confidence_p75': round(np.percentile(confidences, 75), 3),
                'confidence_p95': round(np.percentile(confidences, 95), 3)
            },
            'statistics': {
                'response_time_std': round(df['response_time'].std(), 3),
                'response_time_variance': round(df['response_time'].var(), 3),
                'confidence_std': round(df['confidence'].std(), 3),
                'query_length_avg': round(df['query_length'].mean(), 0),
                'response_length_avg': round(df['response_length'].mean(), 0)
            }
        }
    
    def _getPerformanceBreakdown(self) -> Dict[str, Any]:
        """Get performance breakdown by categories"""
        if not self.query_history:
            return {}
        
        df = self.getMetricsDataFrame()
        
        # Response time categories
        excellent = (df['response_time'] < 1).sum()
        good = ((df['response_time'] >= 1) & (df['response_time'] < 2)).sum()
        acceptable = ((df['response_time'] >= 2) & (df['response_time'] < 5)).sum()
        slow = (df['response_time'] >= 5).sum()
        
        total = len(df)
        
        return {
            'response_time_breakdown': {
                'excellent': {'count': excellent, 'percentage': round(excellent/total*100, 1)},
                'good': {'count': good, 'percentage': round(good/total*100, 1)},
                'acceptable': {'count': acceptable, 'percentage': round(acceptable/total*100, 1)},
                'slow': {'count': slow, 'percentage': round(slow/total*100, 1)}
            },
            'confidence_breakdown': {
                'high': {'count': (df['confidence'] >= 0.8).sum(), 'percentage': round((df['confidence'] >= 0.8).mean()*100, 1)},
                'medium': {'count': ((df['confidence'] >= 0.6) & (df['confidence'] < 0.8)).sum(), 'percentage': round(((df['confidence'] >= 0.6) & (df['confidence'] < 0.8)).mean()*100, 1)},
                'low': {'count': (df['confidence'] < 0.6).sum(), 'percentage': round((df['confidence'] < 0.6).mean()*100, 1)}
            }
        }
    
    def exportToExcel(self, query_history: List[Dict]) -> bytes:
        """Export comprehensive analytics to Excel with formatting"""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Sheet 1: Executive Summary
            summary = self._generateExecutiveSummary()
            if summary:
                summary_df = pd.DataFrame([summary])
                summary_df.to_excel(writer, sheet_name='Executive Summary', index=False)
            
            # Sheet 2: Query History
            if query_history:
                history_df = pd.DataFrame(query_history)
                history_df.to_excel(writer, sheet_name='Query History', index=False)
            
            # Sheet 3: Performance Metrics
            if self.query_history:
                metrics_data = []
                for q in self.query_history:
                    metrics_data.append({
                        'Timestamp': q.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                        'Session ID': q.session_id[:8],
                        'Query': q.query[:100],
                        'Response Time (s)': round(q.response_time, 3),
                        'Confidence': round(q.confidence, 3),
                        'Chunks Retrieved': q.chunks_retrieved,
                        'Context Size': q.context_size,
                        'Response Length': q.response_length,
                        'Error': 'Yes' if q.error_occurred else 'No'
                    })
                
                if metrics_data:
                    metrics_df = pd.DataFrame(metrics_data)
                    metrics_df.to_excel(writer, sheet_name='Performance Metrics', index=False)
            
            # Sheet 4: Hourly Statistics
            if self.hourly_stats:
                hourly_data = []
                for hour, times in self.hourly_stats.items():
                    hourly_data.append({
                        'Hour': f'{hour:02d}:00',
                        'Query Count': len(times),
                        'Avg Response Time': round(sum(times)/len(times), 3) if times else 0,
                        'Min Response Time': round(min(times), 3) if times else 0,
                        'Max Response Time': round(max(times), 3) if times else 0
                    })
                
                hourly_df = pd.DataFrame(hourly_data)
                hourly_df.to_excel(writer, sheet_name='Hourly Statistics', index=False)
            
            # Sheet 5: Session Summary
            if self.session_analytics:
                session_data = []
                for sid, session in self.session_analytics.items():
                    duration = (session.end_time - session.start_time).total_seconds() / 60 if session.end_time else 0
                    session_data.append({
                        'Session ID': sid[:8],
                        'Start Time': session.start_time.strftime('%Y-%m-%d %H:%M:%S'),
                        'Duration (min)': round(duration, 1),
                        'Query Count': session.query_count,
                        'Avg Response Time': round(session.total_response_time / max(session.query_count, 1), 3),
                        'Avg Confidence': round(session.avg_confidence, 3)
                    })
                
                session_df = pd.DataFrame(session_data)
                session_df.to_excel(writer, sheet_name='Session Summary', index=False)
            
            # Sheet 6: Insights & Recommendations
            insights = self._generateInsights()
            recommendations = self._generateRecommendations()
            max_len = max(len(insights), len(recommendations))
            insights.extend([''] * (max_len - len(insights)))
            recommendations.extend([''] * (max_len - len(recommendations)))
            
            insights_df = pd.DataFrame({
                'Insights': insights,
                'Recommendations': recommendations
            })
            insights_df.to_excel(writer, sheet_name='Insights', index=False)
            
            # Format workbook
            workbook = writer.book
            self._formatExcelWorkbook(workbook)
        
        output.seek(0)
        return output.read()
    
    def _formatExcelWorkbook(self, workbook):
        """Apply formatting to Excel workbook"""
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet in workbook.worksheets:
            # Format headers
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(50, max(12, max_length + 2))
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze top row
            sheet.freeze_panes = 'A2'
